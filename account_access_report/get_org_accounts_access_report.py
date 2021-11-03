import boto3
import botocore
import datetime
import botocore.exceptions
## You need pandas to simplify export of the org account report
import pandas as pd
## Yo need openpyxl to manipulate Excel file
import openpyxl
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo

## Get all the full account path
def get_org_path(parentPath,parentId):
    answers = []
    ## get just accounts
    accounts = orgclient.list_children(
        ParentId = parentId,
        ChildType = "ACCOUNT"
    )
    ## take the accounts, just add the parentpath to the front of it
    for account in accounts["Children"]:
        answers = answers + ["/".join([parentPath,account["Id"]])]
    ## get the OUs
    ous = orgclient.list_children(
        ParentId = parentId,
        ChildType = "ORGANIZATIONAL_UNIT"
    )       
    ## take the OU and pass it back into function and repeat until we get to accounts
    for ou in ous["Children"]:
        answers = answers + get_org_path("/".join([parentPath,ou["Id"]]),ou["Id"])
    return answers
 
## Update timezone, Excel no like timezone
def update_timezone(details):
    new_details = []
    for detail in details:
        if (detail.get("LastAuthenticatedTime",None) is not None):
            detail.update({"LastAuthenticatedTime":(detail["LastAuthenticatedTime"]).replace(tzinfo=None)})
        new_details = new_details + [detail]
    return new_details
       
## This is the bulk of this script, to generate and retrieve access report
def get_org_access_report(job_id,accountname,filename):
    ## table name must begin with a letter or underscore
    table_name = "T" + accountname
    total_details = []
    is_trunk = False
    jobStat = "IN_PROGRESS"
    ## keep running until it's no longer "in_progress"
    while(jobStat == "IN_PROGRESS"):
        response = iamClient.get_organizations_access_report(
            JobId=job_id,
            MaxItems=10,
            SortKey='LAST_AUTHENTICATED_TIME_ASCENDING'
        )
        jobStat = response["JobStatus"]
    ## if it's completed, it's good
    if(jobStat == "COMPLETED"):
        marker = response["Marker"]
        is_trunk = response["IsTruncated"]
        details = update_timezone(response["AccessDetails"])
        total_details = total_details + details
        ## keep running if it's truncated
        while is_trunk:
            response = iamClient.get_organizations_access_report(
                JobId=job_id,
                MaxItems=100,
                SortKey='LAST_AUTHENTICATED_TIME_ASCENDING',
                Marker = marker
            )
            is_trunk = response["IsTruncated"]
            marker = response.get("Marker","")
            detail = response.get("AccessDetails",[])
            ##panda needs timezone unaware datetime stamp to send to excel
            details = update_timezone(response["AccessDetails"])
            total_details = total_details + details
        ## create dataframe
        df = pd.DataFrame(total_details)
        ## Need ExcelWriter so we can write multiple sheets to excel
        ## if_sheet_exists is set to none so we replace
        ## https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.to_excel.html
        with pd.ExcelWriter(filename,mode='a') as writer:  
            df.to_excel(writer, sheet_name=accountname, index=False, header=True)
        
        ## this part is to create a formatted table inside excel
        ## load the excel file
        wb = load_workbook(filename = filename)
        ## load the sheet within the excel 
        sheet_ranges = wb[accountname]
        ## create a range of the cells within that sheet
        table_ref = "A1" + ":" + chr(sheet_ranges.max_column + 64) + str(sheet_ranges.max_row)
        ## define a formatted table
        tab = Table(displayName=table_name, ref=table_ref)
        ## format the formatted table
        style = TableStyleInfo(name="TableStyleLight9")
        tab.tableStyleInfo = style
        ## add the formatted table to the excel
        sheet_ranges.add_table(tab)
        ## update the excel file
        wb.save(filename)

        print("Sent to file")

    else:
        print("Failed to create report")    

######################################################################################################
## MAIN ##############################################################################################
######################################################################################################
## Update the profile to match the credential in your ~/.aws/credentials file
profile = "master"
session = boto3.Session(profile_name=profile)
## Update the region to match the region where you are. IAM and Org are region-free, so it shouldn't matter
region = "us-east-1"
## Using IAM to run the report
iamClient = session.client("iam")
## Using Org to get ou and account ids
orgclient = session.client('organizations')
## get current time
nowmeow = datetime.datetime.now()
## we'll use the timestamp here to uniquely name the file that will be generated
timestamp =nowmeow.strftime("%Y%m%d%H%M")

## Create an empty file, simple way to have a pre-existing excel file
filename = "./" + timestamp + "_" + "org_access_report.xlsx"
print("Generating Report, see ", filename)

empty_df = pd.DataFrame()
empty_df.to_excel(filename)  

## get org info
response = orgclient.describe_organization()
org_id = response["Organization"]["Id"]
response = orgclient.list_roots()
root_id = response["Roots"][0]["Id"]
## This is the starting point
root_path = org_id + "/" + root_id
## get all the paths via this def call
full_paths = get_org_path(root_path, root_id)
## empty list to put jobs
job_content = []
## Pace yourself
## ReportGenerationLimitExceededException: 
### an error occurred (ReportGenerationLimitExceeded) when calling the GenerateOrganizationsAccessReport operation: 
### Maximum number of concurrent jobs exceeded
for entitypath in full_paths:
    print(entitypath, end=":")
    ## generate reports
    response = iamClient.generate_organizations_access_report(
        EntityPath=entitypath
    )
    entitypath_array = entitypath.split("/")
    accountname = entitypath_array[len(entitypath_array)-1]
    ## retrieve reports
    get_org_access_report(response["JobId"],accountname,filename)
    ## store this for later use. Not used at this time. 
    job_content = job_content + [
        {
            "jobid":response["JobId"],
            "entity":entitypath
        }
    ]
